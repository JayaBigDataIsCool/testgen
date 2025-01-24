import os
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Set
import fitz
from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from google import genai
from google.genai.types import GenerateContentConfig
import pandas as pd
from dotenv import load_dotenv
import re
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import uuid
import shutil
import aiofiles

# Constants
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_PAGES = 100
ALLOWED_EXTENSIONS = {".pdf"}

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Initialize FastAPI app
app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")

# Load environment variables
load_dotenv()

# Initialize folders
BASE_UPLOAD_DIR = Path('uploads')
BASE_OUTPUT_DIR = Path('outputs')
BASE_UPLOAD_DIR.mkdir(exist_ok=True)
BASE_OUTPUT_DIR.mkdir(exist_ok=True)

# Update the extraction prompt for better requirement coverage
extraction_prompt = """Analyze this diagram/page in detail and extract ALL testable requirements.
Focus on identifying:

1. Business Rules & Logic
   - Data validation rules
   - Business constraints
   - Calculation logic
   - Processing rules
   - State transitions

2. User Workflows
   - User actions and inputs
   - System responses
   - Navigation paths
   - Error handling
   - Success scenarios

3. Data Requirements
   - Required fields
   - Field validations
   - Data relationships
   - Data transformations

4. System Behaviors
   - Processing steps
   - Integration points
   - Response handling
   - Error conditions
   - Performance criteria

Return ONLY a JSON in this format:
{
    "page_number": <number>,
    "requirements": [
        {
            "req_id": "REQ_P<page_number>_<sequence>",
            "type": "business_rule|workflow|validation|system_behavior",
            "category": "critical|major|normal",
            "description": "detailed requirement description",
            "conditions": [
                "specific condition 1",
                "specific condition 2"
            ],
            "business_rules": [
                "specific rule 1",
                "specific rule 2"
            ],
            "validation_criteria": [
                "specific validation 1",
                "specific validation 2"
            ],
            "expected_behavior": "detailed expected outcome"
        }
    ]
}"""

# Update the test case generation prompt for more comprehensive coverage
test_case_generation_prompt = """Generate detailed test cases for these requirements. For each requirement:

1. Core Functionality Tests
   - Happy path scenarios
   - Alternative flows
   - Edge cases
   - Boundary conditions
   - Performance scenarios

2. Validation Tests
   - Input validation
   - Data format validation
   - Business rule validation
   - Cross-field validations
   - Mandatory field checks

3. Error Handling Tests
   - Invalid inputs
   - System errors
   - Integration failures
   - Timeout scenarios
   - Recovery scenarios

4. Integration Tests
   - API interactions
   - Database operations
   - External system integrations
   - State transitions
   - Data flow validations

5. Non-Functional Tests
   - Performance thresholds
   - Load conditions
   - Security scenarios
   - Accessibility checks
   - Usability aspects

For EACH requirement, generate AT LEAST 12-15 test cases per category above.
Return ONLY a JSON in this format:
{
    "test_cases": [
        {
            "test_id": "TC_<req_id>_<sequence>",
            "requirement_id": "<req_id>",
            "category": "functional|validation|error|integration|non_functional",
            "priority": "high|medium|low",
            "title": "descriptive test case title",
            "preconditions": [
                "precondition 1",
                "precondition 2"
            ],
            "test_steps": [
                "detailed step 1",
                "detailed step 2"
            ],
            "expected_results": [
                "specific expected result 1",
                "specific expected result 2"
            ],
            "test_data": {
                "input": "specific test input data",
                "expected_output": "expected output data"
            },
            "validation_points": [
                "specific validation point 1",
                "specific validation point 2"
            ]
        }
    ]
}"""

def process_page(client: genai.Client, image_path: Path, page_number: int) -> Dict:
    """Process single page and generate test cases immediately"""
    try:
        logger.info(f"Starting processing for page {page_number}")
        
        with open(image_path, 'rb') as img_file:
            img_data = img_file.read()
        
        logger.info(f"Page {page_number}: Extracting requirements...")
        # Extract requirements
        response = client.models.generate_content(
            model="gemini-2.0-flash-exp",
            contents=[{
                "role": "user", 
                "parts": [
                    {"text": extraction_prompt},
                    {"inline_data": {"mime_type": "image/jpeg", "data": img_data}}
                ]
            }],
            config=GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=8000,
                top_p=0.8,
                top_k=40
            )
        )
        
        requirements = extract_json(response.text) if response and response.text else {}
        
        if not requirements.get("requirements"):
            logger.warning(f"Page {page_number}: No requirements extracted")
            return {"page_number": page_number, "requirements": [], "test_cases": []}
            
        logger.info(f"Page {page_number}: Found {len(requirements.get('requirements', []))} requirements")
        
        # Generate test cases immediately for this page
        logger.info(f"Page {page_number}: Generating test cases...")
        test_cases = generate_test_cases(client, requirements, page_number)
        logger.info(f"Page {page_number}: Generated {len(test_cases)} test cases")
        
        result = {
            "page_number": page_number,
            "requirements": requirements.get("requirements", []),
            "test_cases": test_cases
        }
        
        logger.info(f"Page {page_number}: Processing completed")
        return result
        
    except Exception as e:
        logger.error(f"Error processing page {page_number}: {str(e)}", exc_info=True)
        return {"page_number": page_number, "requirements": [], "test_cases": []}

def generate_test_cases(client: genai.Client, requirements: Dict, page_number: int) -> List[Dict]:
    """Generate comprehensive test cases for a single page"""
    try:
        prompt = f"""Based on these requirements:
        {json.dumps(requirements, indent=2)}
        
        Generate HIGH-QUALITY test cases that ensure COMPLETE coverage. Each test case must be:
        1. Business-Focused
           - Validates specific business rules
           - Covers business scenarios
           - Ensures business value
        
        2. Quality-Driven
           - Detailed test steps
           - Specific test data
           - Clear validation points
           - Covers edge cases
        
        Return ONLY a JSON with test cases:
        {{
            "test_cases": [
                {{
                    "id": "TC_P{page_number}_<number>",
                    "priority": "P1|P2|P3",
                    "description": "detailed test description",
                    "requirements_covered": ["REQ_P{page_number}_X"],
                    "preconditions": ["specific condition"],
                    "test_steps": [
                        {{
                            "step": 1,
                            "action": "specific action",
                            "data": "test data",
                            "expected_result": "detailed expected result"
                        }}
                    ],
                    "validation_points": [
                        {{
                            "what": "what to validate",
                            "criteria": "acceptance criteria"
                        }}
                    ]
                }}
            ]
        }}"""

        logger.info(f"Page {page_number}: Calling LLM for test case generation...")
        response = client.models.generate_content(
            model="gemini-2.0-flash-exp",
            contents=[{"role": "user", "parts": [{"text": prompt}]}],
            config=GenerateContentConfig(
                temperature=0.2,
                max_output_tokens=8000,
                top_p=0.9,
                top_k=40
            )
        )
        
        result = extract_json(response.text) if response and response.text else {}
        test_cases = result.get("test_cases", [])
        logger.info(f"Page {page_number}: Test case generation completed")
        return test_cases
        
    except Exception as e:
        logger.error(f"Error generating test cases for page {page_number}: {str(e)}", exc_info=True)
        return []

def extract_json(text: str) -> Dict:
    """Extract JSON from text with robust error handling"""
    try:
        # First try direct JSON parsing
        try:
            return json.loads(text)
        except:
            pass
        
        # Clean up the text
        text = re.sub(r'```json|```|\n|\r', '', text)
        text = text.strip()
        
        # Find JSON content
        json_pattern = r'\{.*\}'
        match = re.search(json_pattern, text, re.DOTALL)
        
        if match:
            json_str = match.group(0)
            return json.loads(json_str)
            
        return {}
            
    except Exception as e:
        logger.error(f"JSON extraction failed: {str(e)}")
        return {}

@app.get("/")
async def read_root():
    """Serve the main page"""
    return FileResponse('static/index.html')

@app.post("/upload")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    """Handle file upload with improved validation and processing"""
    try:
        # Clean up old files first
        cleanup_old_files()
        
        # Generate session ID first
        session_id = str(uuid.uuid4())
        logger.info(f"Starting new upload session: {session_id}")
        
        # Create session folders
        session_upload_dir, session_output_dir = create_session_folders(session_id)
        
        # Validate file extension first
        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS)}"
            )

        # Save file first
        safe_filename = re.sub(r'[^a-zA-Z0-9._-]', '', file.filename)
        file_path = session_upload_dir / safe_filename
        
        file_content = await file.read()
        
        # Validate file size after reading
        if len(file_content) > MAX_FILE_SIZE:
            cleanup_session_files(session_id)
            raise HTTPException(
                status_code=413,
                detail=f"File too large. Maximum size is {MAX_FILE_SIZE/1024/1024}MB"
            )
        
        # Write file
        with open(file_path, 'wb') as f:
            f.write(file_content)
            
        logger.info(f"File saved: {file_path}")
            
        # Start processing in background
        background_tasks.add_task(
            process_document,
            file_path,
            session_id
        )
        
        return JSONResponse({
            "message": "Upload successful. Processing started.",
            "session_id": session_id,
            "status_url": f"/status/{session_id}"
        })
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload error: {str(e)}", exc_info=True)
        if 'session_id' in locals():
            cleanup_session_files(session_id)
        raise HTTPException(status_code=500, detail="Upload failed")

@app.get("/download/{session_id}/excel/{filename}")
async def download_excel(session_id: str, filename: str):
    """Download generated Excel file"""
    try:
        file_path = BASE_OUTPUT_DIR / session_id / filename
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Excel file not found")
            
        # Force file download with proper headers
        headers = {
            'Content-Disposition': f'attachment; filename="test_cases.xlsx"'
        }
            
        return FileResponse(
            path=file_path,
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error downloading Excel: {str(e)}")
        raise HTTPException(status_code=500, detail="Download failed")

@app.get("/download/{session_id}/json/{filename}")
async def download_json(session_id: str, filename: str):
    """Download generated JSON file"""
    try:
        file_path = BASE_OUTPUT_DIR / session_id / filename
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="JSON file not found")
            
        # Force file download with proper headers
        headers = {
            'Content-Disposition': f'attachment; filename="test_cases.json"'
        }
            
        return FileResponse(
            path=file_path,
            headers=headers,
            media_type="application/json"
        )
    except Exception as e:
        logger.error(f"Error downloading JSON: {str(e)}")
        raise HTTPException(status_code=500, detail="Download failed")

def init_gemini() -> genai.Client:
    """Initialize Gemini client"""
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("GEMINI_API_KEY environment variable not set")
    return genai.Client(api_key=api_key)

def pdf_to_images(pdf_path: Path, output_dir: Path) -> List[Path]:
    """Convert PDF pages to images"""
    try:
        logger.info(f"Converting PDF to images: {pdf_path}")
        image_paths = []
        
        with fitz.open(str(pdf_path)) as pdf_document:
            for page_number in range(pdf_document.page_count):
                page = pdf_document[page_number]
                zoom = 2
                mat = fitz.Matrix(zoom, zoom)
                pix = page.get_pixmap(matrix=mat, alpha=False)
                
                image_path = output_dir / f"page_{page_number + 1}.jpg"
                pix.save(str(image_path), output="jpeg", jpg_quality=85)
                
                image_paths.append(image_path)
                
        return image_paths
        
    except Exception as e:
        logger.error(f"Error converting PDF to images: {str(e)}")
        raise

def generate_excel(results: List[Dict], output_dir: Path) -> Path:
    """Generate Excel output with test cases and requirements"""
    try:
        excel_path = output_dir / "test_cases.xlsx"
        
        # Prepare test cases data
        test_case_rows = []
        all_requirements = set()  # To collect all requirements
        
        # Group test cases by flow/scenario
        for page in results:
            test_cases = page.get("test_cases", [])
            e2e_flows = group_test_cases_into_flows(test_cases)
            
            for flow in e2e_flows:
                # Extract requirement IDs from all test cases in the flow
                all_req_ids = set()
                for tc in flow:
                    req_ids = tc.get("requirements_covered", [])
                    all_req_ids.update(req_ids)
                    # Collect requirements for the Requirements tab
                    all_requirements.update(req_ids)
                req_ids_str = ", ".join(sorted(all_req_ids)) if all_req_ids else ""
                
                # Rest of the flow processing remains the same
                descriptions = []
                preconditions = set()
                for tc in flow:
                    if tc.get("description"):
                        descriptions.append(tc["description"])
                    if tc.get("preconditions"):
                        preconditions.update(tc["preconditions"])
                
                base_data = {
                    "Application Name": "",
                    "Test Case Priority": flow[0].get("priority", "Low"),
                    "Test Case Status": "Approved",
                    "Test Case Type": "Manual",
                    "Requirement IDs": req_ids_str,
                    "Test Case Name": f"E2E_{flow[0].get('id', '')}",
                    "Description": " | ".join(descriptions),
                    "Pre-Condition": "\n".join(sorted(preconditions))
                }
                
                steps = []
                step_num = 1
                
                for tc in flow:
                    test_steps = tc.get("test_steps", [])
                    if test_steps:
                        for step in test_steps:
                            steps.append({
                                "step_num": step_num,
                                "description": step.get("action", ""),
                                "expected": step.get("expected_result", "")
                            })
                            step_num += 1
                    elif tc.get("validation_points"):
                        for point in tc["validation_points"]:
                            steps.append({
                                "step_num": step_num,
                                "description": point.get("what", ""),
                                "expected": point.get("criteria", "")
                            })
                            step_num += 1
                
                if not steps:
                    steps.append({
                        "step_num": 1,
                        "description": "",
                        "expected": ""
                    })
                
                for step in steps:
                    row_data = base_data.copy()
                    row_data.update({
                        "Step #": step["step_num"],
                        "Step Description": step["description"],
                        "Expected Result": step["expected"]
                    })
                    test_case_rows.append(row_data)
        
        # Prepare requirements data
        req_rows = []
        for req_id in sorted(all_requirements):
            # Find the description and other details for this requirement
            req_details = find_requirement_details(results, req_id)
            req_rows.append({
                "Requirement ID": req_id,
                "Description": req_details.get("description", ""),
                "Type": req_details.get("type", "Functional"),  # Default to Functional
                "Priority": req_details.get("priority", "Medium"),  # Default to Medium
                "Status": "Approved",  # Default status
                "Related Test Cases": find_related_test_cases_for_req(test_case_rows, req_id)
            })
        
        # Create DataFrames
        test_case_columns = [
            "Application Name", "Test Case Priority", "Test Case Status", "Test Case Type",
            "Requirement IDs", "Test Case Name", "Description", "Pre-Condition",
            "Step #", "Step Description", "Expected Result"
        ]
        req_columns = [
            "Requirement ID", "Description", "Type", "Priority", "Status", "Related Test Cases"
        ]
        
        test_cases_df = pd.DataFrame(test_case_rows, columns=test_case_columns)
        requirements_df = pd.DataFrame(req_rows, columns=req_columns)
        
        # Write to Excel with formatting
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Write Test Cases sheet
            test_cases_df.to_excel(writer, sheet_name='Test Cases', index=False)
            format_excel_sheet(writer.sheets['Test Cases'], test_cases_df)
            
            # Write Requirements sheet
            requirements_df.to_excel(writer, sheet_name='Requirements', index=False)
            format_requirements_sheet(writer.sheets['Requirements'], requirements_df)
                
        logger.info(f"Excel file generated with Test Cases and Requirements: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}")
        raise

def format_excel_sheet(sheet, df):
    """Apply formatting to Excel sheet"""
    # Set column widths
    column_widths = {
        "Application Name": 20,
        "Test Case Priority": 15,
        "Test Case Status": 15,
        "Test Case Type": 15,
        "Requirement IDs": 20,
        "Test Case Name": 30,
        "Description": 40,
        "Pre-Condition": 40,
        "Step #": 10,
        "Step Description": 50,
        "Expected Result": 50
    }
    
    # Format headers
    for idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        sheet.column_dimensions[col_letter].width = column_widths.get(col, 30)
        
        header_cell = sheet.cell(row=1, column=idx)
        header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # Format content and merge cells
    current_test_case = None
    merge_start_row = 2
    
    for row in range(2, sheet.max_row + 1):
        test_case = sheet.cell(row=row, column=6).value  # Test Case Name column
        
        # Format all cells in the row
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Handle merging for new test case
        if test_case != current_test_case:
            if current_test_case is not None and merge_start_row < row:
                # Merge cells for previous test case (columns A-H)
                for col in range(1, 9):  # A through H
                    if row - merge_start_row > 1:  # Only merge if there are multiple rows
                        sheet.merge_cells(
                            start_row=merge_start_row,
                            start_column=col,
                            end_row=row - 1,
                            end_column=col
                        )
            merge_start_row = row
            current_test_case = test_case
    
    # Handle merging for the last test case
    if current_test_case is not None and merge_start_row < sheet.max_row:
        for col in range(1, 9):  # A through H
            if sheet.max_row - merge_start_row > 0:  # Only merge if there are multiple rows
                sheet.merge_cells(
                    start_row=merge_start_row,
                    start_column=col,
                    end_row=sheet.max_row,
                    end_column=col
                )

def format_requirements_sheet(sheet, df):
    """Apply formatting to Requirements sheet"""
    # Set column widths
    column_widths = {
        "Requirement ID": 20,
        "Description": 60,
        "Type": 15,
        "Priority": 15,
        "Status": 15,
        "Related Test Cases": 40
    }
    
    # Format headers
    for idx, col in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        sheet.column_dimensions[col_letter].width = column_widths.get(col, 30)
        
        header_cell = sheet.cell(row=1, column=idx)
        header_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_cell.font = Font(color="FFFFFF", bold=True)
        header_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    
    # Format content
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Add subtle alternating row colors
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

def find_requirement_details(results: List[Dict], req_id: str) -> Dict:
    """Find requirement details from the results"""
    for page in results:
        for test_case in page.get("test_cases", []):
            if req_id in test_case.get("requirements_covered", []):
                # Extract requirement details from test case
                return {
                    "description": test_case.get("description", ""),
                    "type": "Functional",  # Default type
                    "priority": test_case.get("priority", "Medium")
                }
    return {}

def find_related_test_cases_for_req(test_case_rows: List[Dict], req_id: str) -> str:
    """Find all test cases related to a requirement"""
    related_cases = set()
    for row in test_case_rows:
        if req_id in row["Requirement IDs"].split(", "):
            related_cases.add(row["Test Case Name"])
    return ", ".join(sorted(related_cases))

def create_session_folders(session_id: str) -> Tuple[Path, Path]:
    """Create session-specific folders for uploads and outputs"""
    try:
        # Create session-specific directories
        session_upload_dir = BASE_UPLOAD_DIR / session_id
        session_output_dir = BASE_OUTPUT_DIR / session_id
        
        # Create directories if they don't exist
        session_upload_dir.mkdir(parents=True, exist_ok=True)
        session_output_dir.mkdir(parents=True, exist_ok=True)
        
        return session_upload_dir, session_output_dir
    except Exception as e:
        logger.error(f"Error creating session folders: {str(e)}")
        raise

def cleanup_old_files():
    """Clean up all files from output and upload directories"""
    try:
        logger.info("Cleaning up old files")
        
        # Clean root output directory
        for item in BASE_OUTPUT_DIR.glob('*'):
            try:
                if item.is_file():
                    item.unlink()
                elif item.is_dir():
                    shutil.rmtree(item)
            except Exception as e:
                logger.warning(f"Failed to delete {item}: {str(e)}")
                
        # Clean root upload directory
        for item in BASE_UPLOAD_DIR.glob('*'):
            try:
                if item.is_file():
                    item.unlink()
                elif item.is_dir():
                    shutil.rmtree(item)
            except Exception as e:
                logger.warning(f"Failed to delete {item}: {str(e)}")
                
        logger.info("Cleanup completed")
    except Exception as e:
        logger.error(f"Error during cleanup: {str(e)}")

def cleanup_session_files(session_id: str):
    """Clean up session-specific files safely"""
    try:
        # Clean up upload directory
        upload_dir = BASE_UPLOAD_DIR / session_id
        if upload_dir.exists():
            for item in upload_dir.glob('*'):
                try:
                    if item.is_file():
                        item.unlink(missing_ok=True)
                except Exception as e:
                    logger.warning(f"Failed to delete file {item}: {str(e)}")
            try:
                upload_dir.rmdir()  # Only remove if empty
            except Exception as e:
                logger.warning(f"Failed to remove upload dir: {str(e)}")
            
        # Clean up output directory
        output_dir = BASE_OUTPUT_DIR / session_id
        if output_dir.exists():
            for item in output_dir.glob('*'):
                try:
                    if item.is_file():
                        item.unlink(missing_ok=True)
                except Exception as e:
                    logger.warning(f"Failed to delete file {item}: {str(e)}")
            try:
                output_dir.rmdir()  # Only remove if empty
            except Exception as e:
                logger.warning(f"Failed to remove output dir: {str(e)}")
            
        logger.info(f"Cleaned up session: {session_id}")
    except Exception as e:
        logger.error(f"Error cleaning up session {session_id}: {str(e)}")

@app.get("/status/{session_id}")
async def get_status(session_id: str):
    """Get processing status for a session"""
    try:
        output_dir = BASE_OUTPUT_DIR / session_id
        if not output_dir.exists():
            return JSONResponse({
                "status": "not_found",
                "message": "Session not found"
            })
            
        status_file = output_dir / "processing_status.json"
        if not status_file.exists():
            return JSONResponse({
                "status": "not_found",
                "message": "Status not found"
            })
            
        with open(status_file) as f:
            status = json.load(f)
            
        # Add download URLs if processing is complete
        if status["status"] == "completed":
            excel_path = output_dir / "test_cases.xlsx"
            json_path = output_dir / "results.json"
            
            if excel_path.exists() and json_path.exists():
                status.update({
                    "excel_url": f"/download/{session_id}/excel/test_cases.xlsx",
                    "json_url": f"/download/{session_id}/json/results.json"
                })
                
        return JSONResponse(status)
        
    except Exception as e:
        logger.error(f"Error getting status: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    logger.error(f"Global error handler caught: {str(exc)}", exc_info=True)
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal server error. Please try again later."}
    )

async def process_document(
    file_path: Path,
    session_id: str
) -> None:
    """Process document and generate test cases"""
    try:
        output_dir = BASE_OUTPUT_DIR / session_id
        status_file = output_dir / "processing_status.json"
        results_file = output_dir / "results.json"  # Simplified filename
        
        def update_status(status: str, message: str, progress: int = 0):
            with open(status_file, 'w') as f:
                json.dump({
                    "status": status,
                    "message": message,
                    "progress": progress,
                    "session_id": session_id
                }, f)
        
        update_status("processing", "Starting document processing...", 0)
        
        # Convert PDF to images
        update_status("processing", "Converting PDF to images...", 10)
        image_paths = pdf_to_images(file_path, output_dir)
        
        if len(image_paths) > MAX_PAGES:
            raise ValueError(f"Document has too many pages (max: {MAX_PAGES})")
            
        # Initialize Gemini
        update_status("processing", "Initializing AI model...", 20)
        client = init_gemini()
        
        # Process pages
        results = []
        total_pages = len(image_paths)
        
        for idx, image_path in enumerate(image_paths, 1):
            try:
                progress = 20 + (60 * idx // total_pages)  # Progress from 20% to 80%
                update_status(
                    "processing", 
                    f"Processing page {idx}/{total_pages}...",
                    progress
                )
                
                result = process_page(client, image_path, idx)
                results.append(result)
            except Exception as e:
                logger.error(f"Error processing page {idx}: {str(e)}")
                results.append({
                    "page_number": idx,
                    "error": str(e),
                    "requirements": [],
                    "test_cases": []
                })
        
        # Save results with consistent filename
        logger.info("Saving test cases to JSON")
        with open(results_file, 'w') as f:
            json.dump(results, f, indent=2)
            
        # Generate Excel with consistent filename
        logger.info("Generating Excel file")
        excel_path = generate_excel(results, output_dir)
        
        # Update status with file info
        update_status("completed", "Processing completed", 100)
        
        logger.info(f"Processing completed. Files generated in {output_dir}")
        
    except Exception as e:
        logger.error(f"Document processing error: {str(e)}", exc_info=True)
        update_status("failed", f"Processing failed: {str(e)}", 0)
        cleanup_session_files(session_id)

def group_test_cases_into_flows(test_cases: List[Dict]) -> List[List[Dict]]:
    """Group related test cases into E2E flows based on requirements and logic flow"""
    if not test_cases:
        return []
        
    flows = []
    used_cases = set()
    
    for tc in test_cases:
        if tc.get("id") in used_cases:
            continue
            
        # Start a new flow
        current_flow = [tc]
        used_cases.add(tc.get("id"))
        
        # Find related test cases
        related_cases = find_related_test_cases(tc, test_cases, used_cases)
        current_flow.extend(related_cases)
        for related_tc in related_cases:
            used_cases.add(related_tc.get("id"))
        
        flows.append(current_flow)
    
    return flows

def find_related_test_cases(base_case: Dict, all_cases: List[Dict], used_cases: Set[str]) -> List[Dict]:
    """Find test cases related to the base case based on various criteria"""
    related = []
    
    # Get base case attributes for comparison
    base_reqs = set(base_case.get("requirements_covered", []))
    base_id = base_case.get("id", "")
    
    # Extract sequence number if ID follows pattern TC_P{page}_{seq}
    base_seq = int(base_id.split("_")[-1]) if base_id and base_id.split("_")[-1].isdigit() else 0
    
    for tc in all_cases:
        if tc.get("id") in used_cases:
            continue
            
        tc_id = tc.get("id", "")
        tc_reqs = set(tc.get("requirements_covered", []))
        
        # Check if test cases are related based on:
        # 1. Shared requirements
        # 2. Sequential IDs
        # 3. Related descriptions or preconditions
        if (tc_reqs & base_reqs or  # Shared requirements
            (tc_id and tc_id.split("_")[-1].isdigit() and 
             int(tc_id.split("_")[-1]) == base_seq + 1)):  # Sequential ID
            related.append(tc)
    
    return related

if __name__ == "__main__":
    import uvicorn
    logger.info("Starting FastAPI server")
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
